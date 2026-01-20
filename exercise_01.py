import pulp
import openpyxl
from openpyxl.styles import Font, Alignment
import matplotlib.pyplot as plt

# =========================
# OPTIMIZATION ENGINE
# =========================
def execute_production_plan(periods_and_demand, initial_stock, final_stock, costs):
    # Extrai nomes dos períodos e demandas em listas simples
    period_names = [list(p.keys())[0] for p in periods_and_demand]
    demand = [list(p.values())[0] for p in periods_and_demand]
    T = len(demand)

    # Cria um modelo de otimização linear para minimizar custos totais
    model = pulp.LpProblem("ProductionPlan", pulp.LpMinimize)

    # ===============================
    # VARIÁVEIS DE DECISÃO DO MODELO
    # ===============================

    # Produção normal (mesma para todos os períodos)
    N = pulp.LpVariable("normal_prod", lowBound=0)

    # Produção extra, subcontratação e estoque variam por período t
    # extra = [pulp.LpVariable(f"extra_{t}", lowBound=0) for t in range(T)]
    # sub   = [pulp.LpVariable(f"sub_{t}",   lowBound=0) for t in range(T)]
    stock = [pulp.LpVariable(f"stock_{t}", lowBound=0) for t in range(T)]

    extra_blocks = [
        pulp.LpVariable(f"extra_blocks_{t}", lowBound=0, upBound=20, cat="Integer")
        for t in range(T)
    ]
    extra = [50 * extra_blocks[t] for t in range(T)]   # extra real = 50 × blocks

    # SUB: divisível por 80 → blocos inteiros também
    sub_blocks = [
        pulp.LpVariable(f"sub_blocks_{t}", lowBound=0, upBound=20, cat="Integer")
        for t in range(T)
    ]
    sub = [80 * sub_blocks[t] for t in range(T)]  

    # ===============================
    # RESTRIÇÕES DO MODELO
    # ===============================
    for t in range(T):

        # Balanço de estoque:
        # Estoque atual = estoque anterior + produção - demanda
        if t == 0:
            model += stock[t] == initial_stock + N + extra[t] + sub[t] - demand[t]
        else:
            model += stock[t] == stock[t-1] + N + extra[t] + sub[t] - demand[t]

        # Estoque deve ser >= 0 → não permitimos atrasos
        model += stock[t] >= 0

        # Restrição operacional: produção extra + sub não pode ultrapassar 1/3 da produção normal

    # Estoque final desejado
    model += stock[-1] >= final_stock

    # ===============================
    # FUNÇÃO OBJETIVO: CUSTO TOTAL
    # ===============================

    # Custo da produção normal: N ocorre em todos os T períodos
    cost_normal_total = N * T * costs["normal"]

    # Soma dos custos de produção extra e subcontratação por período
    cost_extra_total = pulp.lpSum(extra[t] * costs["extra"] for t in range(T))
    cost_sub_total   = pulp.lpSum(sub[t]   * costs["sub_contract"] for t in range(T))

    # Custo de estoque considerando média entre início e fim do período
    cost_stock_total = pulp.lpSum(
        ((stock[t] + (stock[t-1] if t > 0 else initial_stock)) / 2) * costs["stock"]
        for t in range(T)
    )

    # Função objetivo = soma de todos os custos
    model += cost_normal_total + cost_extra_total + cost_sub_total + cost_stock_total

    # Resolve o modelo usando o solver padrão do PuLP
    model.solve(pulp.PULP_CBC_CMD(msg=False))
    print(pulp.LpStatus[model.status])

    # ===============================
    # COLETA DE RESULTADOS
    # ===============================
    result = {
        "periods": period_names,
        "demand": demand,
        "normal": [N.value()] * T,
        "extra": [extra[t].value() for t in range(T)],
        "sub":   [sub[t].value()   for t in range(T)],
        "stock_final": [stock[t].value() for t in range(T)],
        "stock_initial": [],
        "stock_avg": [],
        "prod_total": [],
        "prod_minus_dem": [],
        "delays": [0] * T,
        "costs": {},
        "total_cost": pulp.value(model.objective)
    }

    # Reconstrução do estoque inicial e médio por período
    si = initial_stock
    for t in range(T):
        result["stock_initial"].append(si)
        sf = result["stock_final"][t]
        result["stock_avg"].append((si + sf) / 2)
        si = sf

    # Cálculo da produção total e produção - demanda
    total_prod = [
        result["normal"][t] + result["extra"][t] + result["sub"][t]
        for t in range(T)
    ]
    result["prod_total"] = total_prod
    result["prod_minus_dem"] = [total_prod[t] - demand[t] for t in range(T)]

    # Cálculo dos custos por período para exportar ao Excel
    result["costs"]["normal_unit"] = costs["normal"]
    result["costs"]["extra_unit"]  = costs["extra"]
    result["costs"]["sub_unit"]    = costs["sub_contract"]
    result["costs"]["stock_unit"]  = costs["stock"]
    result["costs"]["delay_unit"]  = costs["delay"]

    result["costs"]["normal"] = [result["normal"][t]    * costs["normal"]      for t in range(T)]
    result["costs"]["extra"]  = [result["extra"][t]     * costs["extra"]       for t in range(T)]
    result["costs"]["sub"]    = [result["sub"][t]       * costs["sub_contract"] for t in range(T)]
    result["costs"]["stock"]  = [result["stock_avg"][t] * costs["stock"]       for t in range(T)]
    result["costs"]["delay"]  = [0 for _ in range(T)]

    return result


# =========================
# CAPACITY (DEPARTMENTS)
# =========================
def compute_department_capacity(result, times_per_dept):
    """
    times_per_dept: horas por unidade em cada departamento
    Ex: [0.3, 0.2, 0.6, 0.4]
    """

    # Períodos e produção total por período
    periods = result["periods"]
    T = len(periods)
    prod = result["prod_total"]

    # Número de departamentos (ex: 4)
    n_depts = len(times_per_dept)

    # Matriz de horas: dept × período
    hours = [[0.0]*T for _ in range(n_depts)]

    # Total de horas por departamento no ano
    totals = [0.0]*n_depts

    # Calcula horas requeridas: horas = produção * tempo/unit
    for d in range(n_depts):
        for t in range(T):
            hours[d][t] = prod[t] * times_per_dept[d]
            totals[d] += hours[d][t]

    return {
        "times_per_dept": times_per_dept,
        "hours": hours,
        "totals": totals,
        "periods": periods
    }


# =========================
# EXPORT TO XLSX (with TOTAL column and departments table)
# =========================
def export_to_xlsx(result, dept_capacity=None, filename="plano_producao_completo.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plano"

    periods = result["periods"]
    T = len(periods)

    # Header
    header = ["Item"] + periods + ["TOTAL"]
    ws.append(header)
    bold = Font(bold=True)
    for c in range(1, len(header)+1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    row_idx = 2

    def write_values(label, values):
        nonlocal row_idx
        ws.cell(row=row_idx, column=1, value=label)
        for i, v in enumerate(values):
            ws.cell(row=row_idx, column=2 + i, value=v)
        # TOTAL column: sum across periods using formula
        first_col = 2
        last_col = 2 + T - 1
        ws.cell(row=row_idx, column=2 + T, value=f"=SUM(B{row_idx}:{openpyxl.utils.get_column_letter(last_col)}{row_idx})")
        row_idx += 1

    # Demand
    write_values("Demanda", result["demand"])

    # Production lines
    write_values("Produção Normal", result["normal"])
    write_values("Turno Extra", result["extra"])
    write_values("Subcontratação", result["sub"])
    write_values("Produção Total", result["prod_total"])
    write_values("Prod - Demanda", result["prod_minus_dem"])

    # Estoques
    write_values("Estoque Inicial", result["stock_initial"])
    write_values("Estoque Final", result["stock_final"])
    write_values("Estoque Médio", result["stock_avg"])
    write_values("Atrasos", result["delays"])

    # Empty row
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Custos $").font = bold
    row_idx += 1

    # Costs rows
    write_values(f"Normal (R$ {result['costs']['normal_unit']})", result["costs"]["normal"])
    write_values(f"Extra (R$ {result['costs']['extra_unit']})", result["costs"]["extra"])
    write_values(f"Subcontratação (R$ {result['costs']['sub_unit']})", result["costs"]["sub"])
    write_values(f"Estoque (R$ {result['costs']['stock_unit']})", result["costs"]["stock"])
    write_values(f"Atrasos (R$ {result['costs']['delay_unit']})", result["costs"]["delay"])

    # Row for total costs per period (sum vertically)
    # compute column letters
    first_cost_row = row_idx - 5  # where "Normal(cost)" was written
    last_cost_row = row_idx - 1
    ws.cell(row=row_idx, column=1, value="TOTAL CUSTOS POR PERÍODO").font = bold
    for i in range(T):
        col_letter = openpyxl.utils.get_column_letter(2 + i)
        ws.cell(row=row_idx, column=2 + i, value=f"=SUM({col_letter}{first_cost_row}:{col_letter}{last_cost_row})")
    # total grand
    ws.cell(row=row_idx, column=2 + T, value=f"=SUM(B{row_idx}:{openpyxl.utils.get_column_letter(2+T-1)}{row_idx})")
    row_idx += 2

    # Department capacity table (if provided)
    if dept_capacity is not None:
        ws.cell(row=row_idx, column=1, value="Capacidade necessária por departamento (horas)").font = bold
        row_idx += 1
        # header
        ws.cell(row=row_idx, column=1, value="Departamento").font = bold
        for i, p in enumerate(dept_capacity["periods"]):
            ws.cell(row=row_idx, column=2 + i, value=p).font = bold
        ws.cell(row=row_idx, column=2 + T, value="TOTAL").font = bold
        row_idx += 1
        # each department
        for d_idx, times in enumerate(dept_capacity["times_per_dept"]):
            ws.cell(row=row_idx, column=1, value=f"Dep {d_idx+1} (h/un={times})")
            for t in range(T):
                ws.cell(row=row_idx, column=2 + t, value=round(dept_capacity["hours"][d_idx][t], 3))
            # total for dept
            ws.cell(row=row_idx, column=2 + T, value=round(dept_capacity["totals"][d_idx], 3))
            row_idx += 1
        # row with totals per period (sum of depts)
        ws.cell(row=row_idx, column=1, value="Total horas por período").font = bold
        for t in range(T):
            # sum over dept rows just written: dept rows start at (row_idx - n_depts)
            start = row_idx - len(dept_capacity["times_per_dept"])
            col_letter = openpyxl.utils.get_column_letter(2 + t)
            ws.cell(row=row_idx, column=2 + t, value=f"=SUM({col_letter}{start}:{col_letter}{start+len(dept_capacity['times_per_dept'])-1})")
        # grand total
        ws.cell(row=row_idx, column=2 + T, value=f"=SUM(B{row_idx}:{openpyxl.utils.get_column_letter(1+T)}{row_idx})")

    # Adjust column widths
    ws.column_dimensions['A'].width = 36
    for c in range(2, 3 + T):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 14

    wb.save(filename)
    return filename

# =========================
# GENERATE GRAPH
# =========================
def generate_graph(result, filename="grafico_plano.png"):
    periods = result["periods"]
    demand = result["demand"]
    normal = result["normal"]
    total_prod = result["prod_total"]
    stock = result["stock_final"]

    plt.figure(figsize=(10, 5))
    plt.plot(periods, demand, label="Demanda", marker='o')
    plt.plot(periods, normal, label="Produção Normal", marker='o')
    plt.plot(periods, total_prod, label="Produção Total", marker='o')
    plt.plot(periods, stock, label="Estoque Final", marker='o')

    plt.legend()
    plt.xlabel("Período")
    plt.ylabel("Quantidade")
    plt.title("Plano de Produção — Comparativo")

    plt.grid(True)
    plt.savefig(filename, bbox_inches="tight")
    plt.close()
    return filename

# =========================
# USO / EXECUÇÃO
# =========================
if __name__ == "__main__":
    periods_and_demand = [
        {"1 trimestre": 28000},
        {"2 trimestre": 25000},
        {"3 trimestre": 11000},
        {"4 trimestre": 10000}
    ]
    initial_stock = 5000
    final_stock = 0
    costs = {"normal":5, "extra":10, "sub_contract":16, "stock":3, "delay":160}

    # run optimizer
    res = execute_production_plan(periods_and_demand, initial_stock, final_stock, costs)
    # department times (h per unit) as given in your image/question
    times_per_dept = [0.8, 0.5, 0.6, 0.6]
    dept_cap = compute_department_capacity(res, times_per_dept)

    # save xlsx and graph
    xlsx_path = export_to_xlsx(res, dept_capacity=dept_cap, filename="test_exercise.xlsx")
    graph_path = generate_graph(res, filename="test_exercise.png")

    print("Resultado (res):")
    print(res)
    print("Arquivo XLSX gerado em:", xlsx_path)
    print("Gráfico gerado em:", graph_path)